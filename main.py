import openpyxl 

inv_file = openpyxl.load_workbook("inventory.xlsx") #this will give us whole file content
product_info = inv_file["Sheet1"] # information of sheet 1 is in variabnle whole_inoformtio


#first task is to count number of the companies and their total count, and storew it in the Dict


product_per_supplier= {}

total_inventory = {}

info_less_ten = {}

#product per supplier
for cursor in range(2, product_info.max_row+1): #by default the max_row starts with 0, and so it will be 0 to 74, so we are adding +1
    supplier_name = product_info.cell(cursor, 4).value #row second, column 4 (2,4)    
    if supplier_name in product_per_supplier:
        count_supplier_name = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = count_supplier_name+1  # +1 in total, so i need the count 
    else:
        #print("New Supplie added")
        product_per_supplier[supplier_name]=1

#total inventory 
    
    if supplier_name in total_inventory:
        
        previous_value= total_inventory[supplier_name]           
        total_inventory[supplier_name] = previous_value+(product_info.cell(cursor, 2).value*product_info.cell(cursor,3).value) 

    else:
        print("New value added")
        total_inventory[supplier_name] = product_info.cell(cursor, 2).value*product_info.cell(cursor,3).value 


#products with inventory less 10 

    
    if product_info.cell(cursor, 2).value < 10:
        info_less_ten[int(product_info.cell(cursor,1).value)] = int(product_info.cell(cursor, 2).value) #bty defualt the values are interpreted as float numbers 

#new column , 5th column with the price*inventory 

    product_info.cell(cursor, 5).value = product_info.cell(cursor, 2).value * product_info.cell(cursor,3).value
    
inv_file.save("inv_updated.xlsx")


print(product_per_supplier)
print(total_inventory)
print(info_less_ten)